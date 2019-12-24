import windnd
import time
import sys
import tkinter as tk
import ctypes
from EHLLAPI import Emulator
import openpyxl as opxl
from openpyxl.styles import  PatternFill
#print(hapi.set_cursor(753))
#print(hapi.get_field(453,20))
#print(hapi.get_cursor())

# Load DLL into memory.
hapi = Emulator()
#窗口
root = tk.Tk()
root.geometry("300x200")

#数据存储     B          C         E          F           I             J             L
dataList = {"Cust":[],"Deli":[],"MPN":[],"PO_num":[],"U_price":[],"Deli_date":[],"Po_qty":[]}
errorList =[]

#加载数据------------------------------
filePath =""
excel_wb={}
def loadExcel_Data():
    global excel_wb
    lable1['text'] = "开始加载EXCEL数据..."
    #Column (B=2 C=3 E=5 F=6 I=9 J=10 L=12)
    excel_wb = opxl.load_workbook(filePath,data_only=True)
    ws = excel_wb.active

    selectlist = [2,3,5,6,9,10,12] #指定获取某列数据
    row_range=ws[2:ws.max_row]

    #每获取一行数据存储在临时tempArr列表里
    tempArr=[]
    for row in row_range:
        for cell in row:
            if(cell.column in selectlist):
                tempArr.append(cell.value)
        count = 0
        #tempArr列表存储的一行数据依次存入datList里每个key
        for key in dataList:
            dataList[key].append(tempArr[count])
            count = count+1
        tempArr=[]

    #处理Deli_date 数据格式
    formatDate_Arr = []
    for item in dataList["Deli_date"]:
        dt = item.date()
        dateYear = str(dt.year)[2:4] #只取年份后面两位
        dateMonth = str(dt.month)
        dateDay = str(dt.day)

        if(len(dateDay) <2 ):
            dateDay = "0" + dateDay
       
        if(len(dateMonth) <2):
            dateMonth = "0" + dateMonth
        #格式化日月年
        formatDate_Arr.append(dateDay+dateMonth+dateYear) 
    dataList["Deli_date"] = formatDate_Arr
    print(dataList)
    formatDate_Arr = [] 
    lable1['text'] = "数据加载完毕！"
    button['state'] = 'normal'

#检测feild 取到得得价格是否和传入的U_Price 价格一致，一致返回True否则False
def check_uPrice(up):
    uprice = up
    sys_unit_price_str = str(hapi.get_field(664,14))
    sys_unit_price = float(sys_unit_price_str.split('u')[0][2:len(sys_unit_price_str)])
    if(sys_unit_price == uprice ):
        return True
    else:
        return False

def search_str(expect_str):
    try:
        hapi.search_str(expect_str,0)
        return True
    except:
        return False

def wait_screen(es):
    print("检测开始{}".format(es))
    #count_s = 0
    while search_str(es) == False:
        time.sleep(0.2)
        print("wait for expect string...")
        #超时退程序
        #count_s = count_s + 0.2
        # if(count_s > 10):
        #   sys.exit("Time out!")
    print("Found string!")

def processFile():
    if(hapi.connect()==0):
        print("Connected!")

        #锁定PS防止其它程序输入

        hapi.lock_kb()

        #进入操作流程       
        for index,item in enumerate(dataList['U_price']):
             
            print("正在处理第{}条".format(index+1))
            #screen1
            
            hapi.copy_str_to_field(str(dataList["Cust"][index]),173)
            hapi.send_keys("@T")
            hapi.copy_str_to_field(str(dataList["Deli"][index]),253)
            hapi.send_keys("@T")
            hapi.copy_str_to_field(str(dataList["MPN"][index]),333)
            hapi.send_keys("@T")
            hapi.copy_str_to_field(str(dataList["PO_num"][index]),359)
            hapi.send_keys("@E@E")
            wait_screen("CURRENCY")

            # #screen2
            # #检测单价是否一致不一致塞到errorList 里并且返回screen1
            if(check_uPrice(item)==False):
                errorList.append(index+2)
                hapi.send_keys("@c@c")
                wait_screen("F7:ITEM")
                continue

            hapi.copy_str_to_field(str(dataList["Deli_date"][index]),645)
            hapi.set_cursor(651)
            hapi.send_keys("@O@O@T")
            hapi.copy_str_to_field(str(dataList["Po_qty"][index]),654)
            hapi.send_keys("@T@E@a")
            wait_screen("F7:ITEM")
            print("第{}条处理完毕".format(index+1))
            
        if(len(errorList)==0):
            lable1['text'] = "处理完毕!"
        else:
            lable1['text'] = "处理结束,{}条异常!!!".format(len(errorList)) 
        
    #将异常元素的EXCEL单元格用红色标记
    for item in errorList:
        ws1 = excel_wb.active
        ws1["I{}".format(item)].fill = PatternFill("solid", fgColor="FF0000")
        #print(filePath)
    excel_wb.save(filePath)

    #完成后关闭API连接
    if(hapi.disconnect()==0):
        print("Disconnected")
#窗口布局
lable = tk.Label(root,bg="pink",width=55,height=5,text="文件拖放到这里")
lable1 = tk.Label(root,bg="orange",width=55,height=1,text="未加载文件")
button = tk.Button(root,text="开始录入",state=tk.DISABLED,command=processFile)

# 获取文件路径后加载文件数据
def func(ls):
        global filePath
        filePath = str(ls)[3:-2]
        showText = filePath.split("\\")[-1]
        lable['text'] = showText
        loadExcel_Data()
# windows 挂钩
windnd.hook_dropfiles(lable.winfo_id(),func)
lable.pack()
lable1.pack()
button.pack()


root.mainloop()
