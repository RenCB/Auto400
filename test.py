import openpyxl as opxl

dataList = {"Cust":[],"Deli":[],"MPN":[],"PO":[]}

wb = opxl.load_workbook("C:\\Users\\rencaibao\\Desktop\\MEKT03-321600-PO-1219 -TEST.xlsx",data_only=True)

ws = wb.active

dataList["Cust"].append(ws["B2"].value)
dataList["Deli"].append(ws["C2"].value)
dataList["MPN"].append(ws["E2"].value)
dataList["PO"].append(ws["F2"].value)

print(dataList["Cust"][0])